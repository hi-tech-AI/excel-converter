import pandas as pd
from datetime import datetime
from dateutil import parser
from openpyxl import Workbook

header_list = ["Date", "Time", "Symbol", "Quantity", "Price", "Commission", "Action"]
wb = Workbook()
sheet = wb.active

for item in range(len(header_list)):
    sheet.cell(row=1, column=item + 1).value = header_list[item]

def clean_date_format(date_string):
    if "as of" in date_string:
        date_string = date_string.split(" as of ")[1].replace(' ', '')
    
    formats = [
        "%Y-%m-%d",        # e.g., 2024-07-31
        "%d-%m-%Y",        # e.g., 31-07-2024
        "%m/%d/%Y",        # e.g., 12/27/2024
        "%Y/%m/%d",        # e.g., 2024/07/31
        "%d/%m/%Y"         # e.g., 31/07/2024
    ]
    
    for fmt in formats:
        try:
            date_object = datetime.strptime(date_string, fmt)
            return date_object.strftime("%m/%d/%Y")
        except:
            continue
    
    # Handle ISO 8601 format
    try:
        date_object = datetime.fromisoformat(date_string)
        return date_object.strftime("%m/%d/%Y")
    except:
        pass
    
    # Handle other ISO 8601 formats using dateutil.parser
    try:
        date_object = parser.isoparse(date_string)
        return date_object.strftime("%m/%d/%Y")
    except:
        pass
    
def clean_time_format(time_string):
    try:
        input_format = "%m/%d/%Y %H:%M:%S:%f"
        date_object = datetime.strptime(time_string, input_format)
        formatted_time = date_object.strftime("%H:%M:%S").upper()
        return formatted_time
    except:
        pass

    try:
        parsed_time = parser.parse(time_string)
        formatted_time = parsed_time.strftime("%H:%M:%S")
        return formatted_time
    except ValueError:
        return "Invalid time format"

def clean_date_format_symbol(date_string):
    if "as of" in date_string:
        date_string = date_string.split(" as of ")[1].replace(' ', '')
    
    formats = [
        "%Y-%m-%d",        # e.g., 2024-07-31
        "%d-%m-%Y",        # e.g., 31-07-2024
        "%m/%d/%Y",        # e.g., 12/27/2024
        "%Y/%m/%d",        # e.g., 2024/07/31
        "%d/%m/%Y"         # e.g., 31/07/2024
    ]
    
    for fmt in formats:
        try:
            date_object = datetime.strptime(date_string, fmt)
            return date_object.strftime('%d %b %y')
        except:
            continue
    
    # Handle ISO 8601 format
    try:
        date_object = datetime.fromisoformat(date_string)
        return date_object.strftime('%d %b %y')
    except:
        pass
    
    # Handle other ISO 8601 formats using dateutil.parser
    try:
        date_object = parser.isoparse(date_string)
        return date_object.strftime('%d %b %y')
    except:
        pass

def clean_symbol_format(symbol_string):
    if symbol_string == 'nan':
        return ''

    if len(symbol_string.split(" ")) != 4:
        return symbol_string.upper()
    else:
        symbol = symbol_string.split(" ")[0].upper()

        symbol_date = clean_date_format_symbol(symbol_string.split(" ")[1])
        
        if "$" in symbol_string.split(" ")[2]:
            strike_price = symbol_string.split(" ")[2].replace(",", "")
        else:
            strike_price = "$" + symbol_string.split(" ")[2].replace(",", "")
        
        if symbol_string.split(" ")[-1] == "C":
            type = "Call"
        else:
            type = "Put"
        
        return " ".join([symbol, symbol_date, strike_price, type])

def clean_quantity_format(quantity_string):
    if quantity_string == 'nan':
        return ''
    try:
        quantity_float = float(quantity_string.replace(',', ''))
        return str(quantity_float)
    except ValueError:
        return None

def clean_price_format(price_string):
    if price_string == 'nan':
        return ''
    
    remove_currency = price_string.replace("$", "").replace(",", "").replace(" ", "")
    return f"{float(remove_currency):.2f}"

def clean_commission_format(commission_string):
    if commission_string == 'nan':
        return ''
    
    remove_currency = commission_string.replace("$", "").replace(",", "").replace(" ", "")
    return f"{float(remove_currency):.2f}"

def clean_action_format(action_string):
    if action_string == 'nan':
        return ''
    
    if 'buy' in action_string or 'bought' in action_string or 'Buy' in action_string or 'Bought' in action_string or 'BOUGHT' in action_string:
        return 'Buy'
    elif 'sell' in action_string or 'sold' in action_string or 'Sell' in action_string or 'Sold' in action_string or 'SOLD' in action_string:
        return 'Sell'

def extract_table_from_excel(file_path):
    skip_num = 0
    header_to_skip = True

    while header_to_skip:
        if file_path.split('.')[-1] == "csv":
            df = pd.read_csv(file_path, skiprows=skip_num)
        elif file_path.split('.')[-1] == "xlsx":
            df = pd.read_excel(file_path, skiprows=skip_num)
        else:
            return "Invalid file type"

        header_to_skip = False
        for col in df.columns:
            if "For Account" in col or "Unnamed:" in col:
                header_to_skip = True
                break
        
        skip_num += 1

    if file_path.split('.')[-1] == "csv":
        df = pd.read_csv(file_path, skiprows=skip_num - 1)
    elif file_path.split('.')[-1] == "xlsx":
        df = pd.read_excel(file_path, skiprows=skip_num - 1)

    # Detect non-empty cells
    not_empty_cells = df.notnull().astype(int)
    
    potential_tables = []
    current_table = None

    for row_index, row in not_empty_cells.iterrows():
        row_sum = row.sum()
        if row_sum > 0 and current_table is None:
            # Start of a new table
            current_table = {"start_row": row_index, "end_row": row_index}
        elif row_sum == 0 and current_table is not None:
            # End of the current table
            current_table["end_row"] = row_index - 1
            potential_tables.append(current_table)
            current_table = None
    
    # Handle the case where the last row contains non-empty cells
    if current_table is not None and current_table not in potential_tables:
        current_table["end_row"] = len(not_empty_cells) - 1
        potential_tables.append(current_table)

    # Output the detected tables
    for i, table in enumerate(potential_tables):
        start_row = table["start_row"]
        end_row = table["end_row"]
        table_df = df.iloc[start_row:end_row + 1]
        print(f"\nDetected Table {i+1}: Rows {start_row} to {end_row}")
    
    return table_df

def process_column_data(data_column, function):
    data_list = []
    for item in data_column:
        data_list.append(function(str(item)))
    return data_list

def complete_column(table_df, column_name, function, column):
    if column_name != '':
        if column_name[-1] == '\n':
            column_name = column_name[:-1]
            if len(table_df[column_name]) != 0:
                print(f"Processing column: {column_name}")
                data_list = process_column_data(table_df[column_name], function)
                start_row = 2
                for item in data_list:
                    sheet.cell(row=start_row, column=column).value = item
                    start_row += 1
            else:
                print(f"No columns containing {column_name} found")
    else:
        print(f"No found {column_name}")

def generate_final_result(file_path):
    df = pd.read_excel(file_path)
    df_cleaned = df.dropna(subset=['Action'])
    df_cleaned.to_excel(file_path, index=False)